from aiogram import F, types, Bot
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from datetime import datetime
import os
import logging
import asyncio
from concurrent.futures import ThreadPoolExecutor

import openpyxl

from config import ADMIN_ID, WEEKDAYS
from database import Database
from keyboards import get_main_keyboard, get_remove_keyboard
from states import TextOrderState
from utils import (
    get_target_week_dates,
    get_deadline_status,
    get_week_range_display,
    format_date_for_db,
    create_excel_report
)
from cache import cache

executor = ThreadPoolExecutor(max_workers=1)
db = Database()
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

# ==================== КОМАНДЫ ====================

async def cmd_start(message: types.Message):
    """🚀 Старт с подробной информацией"""
    user = message.from_user
    
    asyncio.create_task(register_user_async(user.id, user.username, user.full_name))
    
    target_dates, week_type, _ = get_target_week_dates()
    week_range = get_week_range_display(target_dates)
    deadline_status = get_deadline_status()
    
    await message.answer(
        f"👋 *Добрый день, {user.first_name}!*\n\n"
        f"🍽️ *Система заказа обедов для инструкторов*\n\n"
        f"📅 *Текущий период заказа:* `{week_range}`\n"
        f"└ {week_type}\n"
        f"{deadline_status}\n\n"
        f"📝 *Что нужно делать:*\n"
        f"• Нажать «📝 Новый заказ»\n"
        f"• Ввести ФИО инструктора\n"
        f"• На каждый день ввести **0**, **1** или **2**\n"
        f"• Проверить и подтвердить заказ\n\n"
        f"👇 *Нажмите кнопку «📝 Новый заказ» чтобы начать*",
        parse_mode="Markdown",
        reply_markup=get_main_keyboard(user.id == ADMIN_ID)
    )

async def register_user_async(user_id, username, full_name):
    """Фоновая регистрация"""
    try:
        db.register_employee(user_id, username, full_name)
    except:
        pass

# ==================== ПОДРОБНЫЙ ЗАКАЗ ====================

async def start_order(message: types.Message, state: FSMContext):
    """📝 Начало заказа с подробной информацией"""
    
    target_dates, week_type, _ = get_target_week_dates()
    date_keys = [format_date_for_db(d) for d in target_dates]
    week_range = get_week_range_display(target_dates)
    
    # Предрасчет всех форматов
    week_data = []
    for i, date_key in enumerate(date_keys):
        date_obj = cache.parse_date(date_key)
        week_data.append({
            'key': date_key,
            'day_name': WEEKDAYS[i],
            'display': date_obj.strftime("%d.%m.%Y"),
            'short': date_obj.strftime("%d.%m"),
            'full_date': date_obj.strftime("%d %B %Y"),
            'weekday_full': date_obj.strftime("%A")
        })
    
    await state.update_data(
        date_keys=date_keys,
        week_data=week_data,
        week_range=week_range,
        week_type=week_type,
        current_day=0,
        meals={}
    )
    
    await state.set_state(TextOrderState.waiting_instructor)
    
    await message.answer(
        f"📝 *Оформление нового заказа*\n\n"
        f"📅 *Период заказа:* `{week_range}`\n"
        f"└ {week_type}\n\n"
        f"👤 *Шаг 1 из 8:* Введите ФИО инструктора\n"
        f"└ Пример: *Иванов Иван Иванович*\n"
        f"└ Или: *Петрова Мария*\n\n"
        f"✏️ Напишите ФИО в ответном сообщении:",
        parse_mode="Markdown",
        reply_markup=get_remove_keyboard()
    )

async def process_instructor(message: types.Message, state: FSMContext):
    """Обработка ФИО с переходом к первому дню"""
    instructor = message.text.strip()
    
    if len(instructor) < 5:
        await message.answer(
            "❌ *Слишком короткое ФИО*\n\n"
            "Пожалуйста, введите полное ФИО:\n"
            "└ Пример: *Иванов Иван Иванович*",
            parse_mode="Markdown"
        )
        return
    
    await state.update_data(instructor=instructor)
    await state.set_state(TextOrderState.waiting_quantity)
    
    # Показываем первый день
    data = await state.get_data()
    day_info = data['week_data'][0]
    
    await message.answer(
        f"👤 *Инструктор:* {instructor}\n"
        f"📅 *Период:* {data['week_range']}\n\n"
        f"📝 *Шаг 2 из 8*\n"
        f"📅 *День 1: {day_info['day_name']}* ({day_info['display']})\n\n"
        f"🍽️ *Сколько обедов заказать на этот день?*\n\n"
        f"└ Введите **0** — не заказывать\n"
        f"└ Введите **1** — один обед\n"
        f"└ Введите **2** — два обеда\n\n"
        f"✏️ Напишите цифру (0, 1 или 2):",
        parse_mode="Markdown"
    )

async def ask_next_day(message: types.Message, state: FSMContext):
    """Задаём следующий день с проверкой"""
    data = await state.get_data()
    current_day = data.get('current_day', 0)
    target_dates = data.get('target_dates', [])
    
    # ПРОВЕРКА: не вышли ли за границы
    if current_day >= len(target_dates):
        await show_summary(message, state)
        return
    
    instructor = data.get('instructor', '')
    week_range = data.get('week_range', '')
    
    date_obj = datetime.strptime(target_dates[current_day], "%Y%m%d")
    day_name = WEEKDAYS[current_day]
    date_str = date_obj.strftime("%d.%m.%Y")
    
    # Прогресс бар
    progress = "🟦" * (current_day) + "⬜" * (7 - current_day)
    
    text = (
        f"👤 *Инструктор:* {instructor}\n"
        f"📅 *Период:* {week_range}\n\n"
        f"📊 *Прогресс:* {current_day + 1}/7\n{progress}\n\n"
        f"📅 *День {current_day + 1}: {day_name}* ({date_str})\n\n"
        f"🍽️ Сколько обедов? (0, 1, 2):"
    )
    
    await message.answer(text, parse_mode="Markdown")

async def process_quantity(message: types.Message, state: FSMContext):
    """⚡ Обработка числа с подробным подтверждением"""
    text = message.text.strip()
    
    if not text.isdigit() or text not in ('0', '1', '2'):
        await message.answer(
            "❌ *Неверный ввод*\n\n"
            "Пожалуйста, введите только **0**, **1** или **2**:\n"
            f"└ 0 — не заказывать\n"
            f"└ 1 — один обед\n"
            f"└ 2 — два обеда",
            parse_mode="Markdown"
        )
        return
    
    quantity = int(text)
    data = await state.get_data()
    current_day = data.get('current_day', 0)
    day_info = data['week_data'][current_day]
    
    # Сохраняем выбор
    meals = data.get('meals', {})
    date_key = data['date_keys'][current_day]
    meals[date_key] = quantity
    
    # Показываем подтверждение выбора
    if quantity == 0:
        confirm = f"❌ *Не заказываем* обеды на {day_info['day_name']} ({day_info['short']})"
    elif quantity == 1:
        confirm = f"✅ *1 обед* на {day_info['day_name']} ({day_info['short']})"
    else:
        confirm = f"✅ *2 обеда* на {day_info['day_name']} ({day_info['short']})"
    
    await message.answer(confirm, parse_mode="Markdown")
    
    next_day = current_day + 1
    
    if next_day >= 7:
        # Все дни заполнены - показываем итоги
        await state.update_data(meals=meals)
        await show_summary(message, state)
        return
    
    # Обновляем состояние и показываем следующий день
    await state.update_data(
        meals=meals,
        current_day=next_day
    )
    
    next_day_info = data['week_data'][next_day]
    
    await message.answer(
        f"👤 *Инструктор:* {data['instructor']}\n"
        f"📅 *Период:* {data['week_range']}\n\n"
        f"📝 *Шаг {next_day + 2} из 8*\n"
        f"📅 *День {next_day + 1}: {next_day_info['day_name']}* ({next_day_info['display']})\n\n"
        f"🍽️ *Сколько обедов заказать на этот день?*\n\n"
        f"└ Введите **0** — не заказывать\n"
        f"└ Введите **1** — один обед\n"
        f"└ Введите **2** — два обеда\n\n"
        f"✏️ Напишите цифру (0, 1 или 2):",
        parse_mode="Markdown"
    )

async def show_summary(message: types.Message, state: FSMContext):
    """📋 Подробный показ итогов"""
    data = await state.get_data()
    meals = data.get('meals', {})
    week_data = data.get('week_data', [])
    instructor = data.get('instructor', '')
    week_range = data.get('week_range', '')
    
    print(f"📊 show_summary: meals={meals}")  # Отладка
    
    # Подсчёт итогов
    total = 0
    days_count = 0
    lines = []
    
    for i, day_info in enumerate(week_data):
        qty = meals.get(day_info['key'], 0)
        print(f"   День {i}: {day_info['key']} = {qty}")  # Отладка
        
        if qty > 0:
            total += qty
            days_count += 1
            lines.append(f"✅ *{day_info['day_name']}* ({day_info['short']}): {qty} обед(ов)")
        else:
            lines.append(f"❌ *{day_info['day_name']}* ({day_info['short']}): 0")
    
    # Формируем подробный итог
    text = (
        f"📋 *Проверьте правильность заказа*\n\n"
        f"👤 *Инструктор:* {instructor}\n"
        f"📅 *Период:* {week_range}\n"
        f"📊 *Итого:* {days_count} дней, {total} обедов\n\n"
        f"*Детализация по дням:*\n"
    )
    
    text += "\n".join(lines)
    
    text += "\n\n⚠️ *Проверьте внимательно!*\n"
    text += "После подтверждения заказ будет сохранён."
    
    await state.update_data(total=total, days_count=days_count)
    await state.set_state(TextOrderState.waiting_confirm)
    
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да, всё верно", callback_data="confirm_yes")],
            [InlineKeyboardButton(text="🔄 Заполнить заново", callback_data="confirm_no")],
            [InlineKeyboardButton(text="❌ Отменить заказ", callback_data="cancel")]
        ]
    )
    
    await message.answer(text, parse_mode="Markdown", reply_markup=keyboard)

async def confirm_order(callback: types.CallbackQuery, state: FSMContext):
    """✅ Подтверждение заказа с сохранением в БД"""
    
    if callback.data == "confirm_yes":
        # Получаем данные из состояния
        data = await state.get_data()
        user_id = callback.from_user.id
        instructor = data.get('instructor')
        meals = data.get('meals', {})
        week_range = data.get('week_range', '')
        
        if not meals:
            await callback.answer("❌ Нет данных для сохранения")
            return
        
        # Счетчики для статистики
        saved_count = 0
        total_meals = 0
        saved_details = []
        
        # Сохраняем каждый выбранный день
        for date_key, quantity in meals.items():
            if quantity > 0:  # Сохраняем только положительные значения
                try:
                    # Вызываем метод сохранения в БД
                    db.save_order(
                        user_id=user_id,
                        instructor_name=instructor,
                        date=date_key,
                        quantity=quantity
                    )
                    saved_count += 1
                    total_meals += quantity
                    
                    # Форматируем дату для красивого вывода
                    date_obj = datetime.strptime(date_key, "%Y%m%d")
                    date_str = date_obj.strftime("%d.%m")
                    saved_details.append(f"{date_str}: {quantity}")
                    
                except Exception as e:
                    print(f"❌ Ошибка сохранения: {e}")
        
        # Очищаем состояние
        await state.clear()
        
        # Очищаем кэш БД (если есть такой метод)
        try:
            db.clear_cache()
        except:
            pass
        
        # Формируем красивое сообщение об успехе
        success_text = (
            f"✅ *Заказ успешно подтверждён!*\n\n"
            f"👤 *Инструктор:* {instructor}\n"
            f"📅 *Период:* {week_range}\n"
            f"📊 *Сохранено дней:* {saved_count}\n"
            f"🍱 *Всего обедов:* {total_meals}\n\n"
        )
        
        # Добавляем детали если их немного
        if saved_details and len(saved_details) <= 7:
            success_text += "*Детали:*\n" + "\n".join([f"  • {d}" for d in saved_details])
        
        success_text += f"\n\n✨ Спасибо! Заказ передан администраторам."
        
        # Отправляем подтверждение
        await callback.message.edit_text(
            success_text,
            parse_mode="Markdown"
        )
        
        # Возвращаем в главное меню
        await callback.message.answer(
            "👇 *Главное меню:*",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard(callback.from_user.id == ADMIN_ID)
        )
        
        print(f"✅ Заказ сохранен: {instructor}, {saved_count} дней, {total_meals} обедов")
    
    elif callback.data == "confirm_no":
        # Начать заново
        data = await state.get_data()
        instructor = data.get('instructor', '')
        week_data = data.get('week_data', [])
        
        await state.update_data(current_day=0, meals={})
        await state.set_state(TextOrderState.waiting_quantity)
        
        if week_data:
            day_info = week_data[0]
            await callback.message.edit_text(
                f"🔄 *Начинаем заново*\n\n"
                f"👤 *Инструктор:* {instructor}\n"
                f"📅 *День 1: {day_info['day_name']}* ({day_info['display']})\n\n"
                f"🍽️ Сколько обедов? (0, 1, 2):",
                parse_mode="Markdown"
            )
        else:
            await start_order(callback.message, state)
    
    else:  # cancel
        await state.clear()
        await callback.message.edit_text(
            "❌ *Заказ отменён*\n\nЕсли передумаете, начните новый заказ.",
            parse_mode="Markdown"
        )
        await callback.message.answer(
            "👇 *Главное меню:*",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard(callback.from_user.id == ADMIN_ID)
        )
    
    await callback.answer()

# ==================== ПРОСМОТР ЗАКАЗОВ ====================

async def show_my_orders(message: types.Message):
    """📋 Мои заказы"""
    user_id = message.from_user.id
    
    # Используем обычный метод, не cached
    orders = db.get_user_orders(user_id)  # ← ИЗМЕНЕНО!
    
    if not orders:
        await message.answer(
            "📭 *У вас нет заказов*",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard(user_id == ADMIN_ID)
        )
        return
    
    # Группируем по инструкторам
    instructors = {}
    for instructor_name, date, quantity in orders:
        if instructor_name not in instructors:
            instructors[instructor_name] = []
        instructors[instructor_name].append((date, quantity))
    
    text = "📋 *Ваши заказы*\n\n"
    total_all = 0
    
    for instructor, items in instructors.items():
        text += f"👤 *{instructor}*\n"
        instructor_total = 0
        
        for date, quantity in sorted(items, reverse=True)[:7]:
            date_obj = datetime.strptime(date, "%Y%m%d")
            date_str = date_obj.strftime("%a %d.%m")
            text += f"  • {date_str}: {quantity}\n"
            instructor_total += quantity
            total_all += quantity
        
        text += f"  ✨ Итого: {instructor_total}\n\n"
    
    text += f"📊 *Всего:* {total_all} обедов"
    
    await message.answer(
        text,
        parse_mode="Markdown",
        reply_markup=get_main_keyboard(user_id == ADMIN_ID)
    )

# ==================== АДМИНКА ====================

async def export_to_excel(message: types.Message, bot: Bot):
    """📊 Выгрузить Excel"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ *Доступ запрещён*\n\nЭта команда только для администратора.")
        return
    
    status = await message.answer("🔄 *Формирую отчёт...*\nЭто займёт несколько секунд.")
    
    try:
        # Используем обычный метод, не cached
        all_orders = db.get_all_orders()  # ← ИЗМЕНЕНО!
        
        if not all_orders:
            await status.edit_text("📭 *Нет заказов для выгрузки*")
            return
        
        target_dates, _, _ = get_target_week_dates()
        
        # Создаём Excel отчёт
        temp_path, saved_path = create_excel_report(all_orders, target_dates, save_copy=True)
        
        await message.answer_document(
            types.FSInputFile(temp_path),
            caption=f"📊 *Отчёт по заказам готов*\n💾 Сохранён в папке exports/"
        )
        
        os.remove(temp_path)
        await status.delete()
        
    except Exception as e:
        await status.edit_text(f"❌ *Ошибка:* {str(e)[:50]}")
        logger.error(f"Excel export error: {e}")

async def subscribe_notifications(message: types.Message):
    """🔔 Подписаться на уведомления"""
    user_id = message.from_user.id
    db.subscribe_user(user_id)
    await message.answer(
        "✅ *Вы подписались на уведомления*\n\n"
        "📅 Каждую пятницу в 08:00 я буду напоминать о заказе обедов.",
        parse_mode="Markdown"
    )

async def unsubscribe_notifications(message: types.Message):
    """🔕 Отписаться от уведомлений"""
    user_id = message.from_user.id
    db.unsubscribe_user(user_id)
    await message.answer(
        "❌ *Вы отписались от уведомлений*\n\n"
        "Если захотите снова получать напоминания, нажмите «🔔 Подписаться».",
        parse_mode="Markdown"
    )

async def show_excel_history(message: types.Message):
    """📚 Показать историю Excel отчётов (все листы)"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Доступ запрещён")
        return
    
    import glob
    import os
    from config import EXPORT_PATH
    
    files = glob.glob(os.path.join(EXPORT_PATH, "заказы_архив_*.xlsx"))
    
    if not files:
        await message.answer("📭 Нет сохранённых отчётов")
        return
    
    files.sort(reverse=True)
    
    text = "📚 *Архив Excel отчётов:*\n\n"
    
    for i, file in enumerate(files[:10], 1):
        filename = os.path.basename(file)
        size = os.path.getsize(file) / 1024
        
        # Открываем файл и читаем названия листов
        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            sheets = ", ".join(wb.sheetnames[:3])
            if len(wb.sheetnames) > 3:
                sheets += f" и ещё {len(wb.sheetnames) - 3}"
            wb.close()
        except:
            sheets = "не удалось прочитать"
        
        text += f"{i}. `{filename}`\n"
        text += f"   📊 Листы: {sheets}\n"
        text += f"   📦 {size:.1f} KB\n\n"
    
    await message.answer(text, parse_mode="Markdown")