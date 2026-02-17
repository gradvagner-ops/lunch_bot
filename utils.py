from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import shutil
from config import WEEKDAYS, COMPANY_NAME, EXPORT_PATH

# Константы дедлайна
DEADLINE_DAY = 4  # Пятница
DEADLINE_HOUR = 16
DEADLINE_MINUTE = 0

# ==================== ДАТЫ И ДЕДЛАЙНЫ ====================

def get_target_week_dates():
    """Определяет целевую неделю для заказа"""
    now = datetime.now()
    current_weekday = now.weekday()
    current_hour = now.hour
    current_minute = now.minute
    
    # Проверка дедлайна
    is_after_deadline = False
    
    if current_weekday > DEADLINE_DAY:
        is_after_deadline = True
    elif current_weekday == DEADLINE_DAY:
        if current_hour > DEADLINE_HOUR or (current_hour == DEADLINE_HOUR and current_minute >= DEADLINE_MINUTE):
            is_after_deadline = True
    
    # Рассчитываем целевой понедельник
    days_to_monday = (7 - current_weekday) % 7
    next_monday = now + timedelta(days=days_to_monday)
    
    if is_after_deadline:
        target_monday = next_monday + timedelta(days=7)
        week_type = "через неделю"
    else:
        target_monday = next_monday
        week_type = "следующую неделю"
    
    # Генерируем 7 дней
    dates = []
    for i in range(7):
        dates.append(target_monday + timedelta(days=i))
    
    return dates, week_type, is_after_deadline

def get_deadline_status():
    """Возвращает статус дедлайна для отображения пользователю"""
    now = datetime.now()
    weekday = now.weekday()
    hour = now.hour
    minute = now.minute
    
    # До пятницы
    if weekday < DEADLINE_DAY:
        days_left = DEADLINE_DAY - weekday
        if days_left == 1:
            return f"⏳ Дедлайн: завтра до 16:00"
        else:
            return f"⏳ Дедлайн: пятница 16:00 (осталось {days_left} дн.)"
    
    # Пятница
    elif weekday == DEADLINE_DAY:
        if hour < DEADLINE_HOUR:
            hours_left = DEADLINE_HOUR - hour - 1
            minutes_left = 60 - minute
            return f"⏳ Сегодня до 16:00 (осталось {hours_left} ч {minutes_left} мин)"
        else:
            return "🔓 Приём заказов на неделю через одну"
    
    # Суббота-воскресенье
    else:
        return "🔓 Приём заказов на неделю через одну"


def format_date_for_db(date_obj):
    """Форматирование для БД"""
    return date_obj.strftime("%Y%m%d")

def format_date_for_display(date_obj):
    """Форматирование для показа"""
    return date_obj.strftime("%d.%m.%Y")

def get_week_range_display(dates):
    """Диапазон недели"""
    start = dates[0].strftime("%d.%m")
    end = dates[6].strftime("%d.%m.%Y")
    return f"{start} - {end}"

# ==================== EXCEL ОТЧЁТЫ ====================

def create_excel_report(all_orders, dates, save_copy=True):
    """
    Создаёт Excel файл со всеми заказами
    
    Параметры:
        all_orders: список кортежей (user_id, full_name, instructor_name, date, quantity)
        dates: список дат недели
        save_copy: сохранять ли копию
    
    Возвращает:
        (temp_path, saved_path) - пути к файлам
    """
    
    # Создаём папку для экспорта, если её нет
    os.makedirs(EXPORT_PATH, exist_ok=True)
    
    # Генерируем имя файла
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"заказы_{COMPANY_NAME}_{timestamp}.xlsx"
    temp_path = os.path.join(EXPORT_PATH, f"temp_{filename}")
    saved_path = os.path.join(EXPORT_PATH, filename)
    
    # Создаём книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы на неделю"
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Стиль для итогов
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # ===== ШАПКА ОТЧЁТА =====
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Заказы обедов • {COMPANY_NAME}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # Период
    ws.merge_cells('A2:J2')
    start_date = dates[0].strftime("%d.%m.%Y")
    end_date = dates[6].strftime("%d.%m.%Y")
    ws['A2'] = f"Период заказа: {start_date} - {end_date}"
    ws['A2'].font = Font(size=11)
    ws['A2'].alignment = center_alignment
    
    # Дата создания
    ws.merge_cells('A3:J3')
    creation_time = datetime.now().strftime('%d.%m.%Y %H:%M')
    ws['A3'] = f"Отчёт создан: {creation_time}"
    ws['A3'].font = Font(size=11)
    ws['A3'].alignment = center_alignment
    
    # ===== ЗАГОЛОВКИ ТАБЛИЦЫ =====
    headers = ["№", "Сотрудник", "Инструктор"]
    
    # Добавляем дни недели
    for i, date in enumerate(dates):
        headers.append(f"{WEEKDAYS[i]}\n{date.strftime('%d.%m')}")
    
    headers.append("Всего")
    
    # Применяем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # ===== ГРУППИРУЕМ ЗАКАЗЫ ПО СОТРУДНИКАМ И ИНСТРУКТОРАМ =====
    orders_by_employee = {}
    
    for order in all_orders:
        # Распаковываем 5 значений
        if len(order) == 5:
            user_id, employee_name, instructor_name, date_str, quantity = order
        else:
            print(f"⚠️ Неправильный формат: {order}")
            continue
        
        # Создаём структуру для хранения
        if employee_name not in orders_by_employee:
            orders_by_employee[employee_name] = {}
        
        if instructor_name not in orders_by_employee[employee_name]:
            orders_by_employee[employee_name][instructor_name] = {}
        
        # Сохраняем количество по дате
        orders_by_employee[employee_name][instructor_name][date_str] = quantity
    
    # ===== ЗАПОЛНЯЕМ ДАННЫЕ =====
    row = 5
    employee_counter = 1
    
    # Словарь для подсчёта итогов по дням
    day_totals = {i: 0 for i in range(7)}
    
    for employee_name in sorted(orders_by_employee.keys()):
        instructors = orders_by_employee[employee_name]
        first_row = True
        
        for instructor_name in sorted(instructors.keys()):
            # Номер сотрудника (только для первой строки)
            if first_row:
                ws.cell(row=row, column=1, value=employee_counter)
                first_row = False
            else:
                ws.cell(row=row, column=1, value="")
            
            # ФИО сотрудника
            ws.cell(row=row, column=2, value=employee_name).border = border
            
            # ФИО инструктора
            ws.cell(row=row, column=3, value=instructor_name).border = border
            
            # Заполняем дни
            total_for_instructor = 0
            col = 4
            
            for i, date in enumerate(dates):
                date_key = date.strftime("%Y%m%d")
                quantity = orders_by_employee[employee_name][instructor_name].get(date_key, 0)
                
                cell = ws.cell(row=row, column=col, value=quantity if quantity > 0 else "-")
                cell.alignment = center_alignment
                cell.border = border
                
                if quantity > 0:
                    total_for_instructor += quantity
                    day_totals[i] += quantity
                
                col += 1
            
            # Итого по инструктору
            total_cell = ws.cell(row=row, column=col, value=total_for_instructor)
            total_cell.alignment = center_alignment
            total_cell.border = border
            total_cell.font = Font(bold=True)
            
            row += 1
        
        employee_counter += 1
        # Пустая строка между сотрудниками
        row += 1
    
    # ===== СТРОКА ИТОГОВ =====
    total_row = row
    
    # Подпись
    ws.cell(row=total_row, column=2, value="ИТОГО ПО ДНЯМ:").font = total_font
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).border = border
    
    # Итоги по дням
    col = 4
    grand_total = 0
    for i in range(7):
        cell = ws.cell(row=total_row, column=col, value=day_totals[i])
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_alignment
        cell.border = border
        grand_total += day_totals[i]
        col += 1
    
    # Общий итог
    total_cell = ws.cell(row=total_row, column=col, value=grand_total)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = center_alignment
    total_cell.border = border
    
    # ===== АВТОПОДБОР ШИРИНЫ КОЛОНОК =====
    for col in range(1, 12):
        max_length = 10
        for r in range(1, total_row + 1):
            cell_value = ws.cell(row=r, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 4, 30)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # ===== СОХРАНЕНИЕ =====
    try:
        # Сохраняем временный файл
        wb.save(temp_path)
        print(f"✅ Excel файл создан: {temp_path}")
        
        # Если нужно сохранить копию
        if save_copy:
            shutil.copy2(temp_path, saved_path)
            print(f"📁 Копия сохранена: {saved_path}")
            return temp_path, saved_path
        
        return temp_path, None
        
    except Exception as e:
        print(f"❌ Ошибка при сохранении Excel: {e}")
        raise